#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
별표 아파트명 ↔ 실제(공식) 명칭 별칭 수집 — juso.go.kr 공식 건물명 대조.

목적: 별표는 축약명("이편한세상 시흥장현")인데 실제 통용·공식명은
      "e편한세상 시흥장현 퍼스트베뉴"처럼 더 길어, 실제명으로 검색하면 안 나오는
      문제를 줄인다. juso 공식 건물명을 자동 수집해 '별칭'으로 등록한다.

추측 금지 원칙:
  - juso명이 별표명을 '포함/확장'하거나(같은 단지, 이름만 김),
    또는 juso 지번이 별표/override 지번과 '일치'하면 → 자동확정(O)
  - 그 외(이름도 다르고 지번도 안 맞음) → 검토 대상(엑셀에서 사람이 O/X)

출력:
  data/apt_aliases.csv               자동확정 별칭(빌드에 즉시 반영)
  out/_audit/별표_별칭_검토.xlsx      전체 후보 + O/X (담당자 검토용)
  data/_audit/juso_apt_cache.json    juso 응답 캐시(재실행 시 재조회 안 함)

사용: PYTHONUTF8=1 python fetch_apt_aliases.py
"""
import csv, glob, json, os, re, time, urllib.parse, urllib.request

KEY = "devU01TX0FVVEgyMDI2MDYyMDExMjM0NTExOTQ2NzM="
CACHE = "data/_audit/juso_apt_cache.json"
ALIAS_CSV = "data/apt_aliases.csv"
XLSX = "out/_audit/별표_별칭_검토.xlsx"

JIBUN_IN_ADDR = re.compile(r"시흥시\s+(\S+?[동리])\s+(산?\d+(?:-\d+)?)")
# 아파트가 아닌데 이름 조각이 우연히 포함된 건물(유치원·상사·산업 등)은 별칭에서 제외
NONRES = re.compile(r"유치원|초등학교|중학교|고등학교|어린이집|주민센터|행정복지|"
                    r"상사|산업|교회|성당|병원|의원|시장|주유소|센터$|상가$")


def nn(s):
    """매칭용 정규화: 소문자·공백/아파트/Ⓐ 제거 + e편한세상↔이편한세상 동치."""
    s = re.sub(r"\s+", "", (s or "")).lower()
    s = s.replace("아파트", "").replace("ⓐ", "").replace("ａ", "")
    s = s.replace("e편한세상", "이편한세상").replace("ｅ편한세상", "이편한세상")
    return s


def load_cache():
    if os.path.exists(CACHE):
        return json.load(open(CACHE, encoding="utf-8"))
    return {}


def juso(keyword, cache):
    if keyword in cache:
        return cache[keyword]
    u = "https://www.juso.go.kr/addrlink/addrLinkApi.do?" + urllib.parse.urlencode(
        {"confmKey": KEY, "currentPage": 1, "countPerPage": 20,
         "keyword": keyword, "resultType": "json"})
    try:
        d = json.load(urllib.request.urlopen(u, timeout=20))
        res = d["results"]
        if res["common"]["errorCode"] != "0":
            out = {"err": res["common"]["errorMessage"], "juso": []}
        else:
            out = {"err": "", "juso": [
                {"bdNm": j.get("bdNm", ""), "jibun": j.get("jibunAddr", "")}
                for j in (res["juso"] or [])]}
    except Exception as e:
        out = {"err": str(e), "juso": []}
    cache[keyword] = out
    time.sleep(0.12)
    return out


def main():
    os.makedirs("data/_audit", exist_ok=True)
    os.makedirs("out/_audit", exist_ok=True)
    cache = load_cache()

    # 별표 아파트 목록 (코드, 동, 별표명, 인덱스n)
    idx = json.load(open("out/apartment_index.json", encoding="utf-8"))["list"]
    apts = {}
    for it in idx:
        apts.setdefault((it["code"], it["n"]),
                        {"code": it["code"], "dong": it["dong"],
                         "apt": it["apt"], "idxn": it["n"], "nn": nn(it["apt"])})
    apts = list(apts.values())

    # override 지번 (검증된 지번) — (code, nn아파트) → set(지번)
    known = build_known()    # (code, nn아파트) → set(지번)  [별표 + override]

    print(f"아파트 {len(apts)}개 juso 대조 시작 (캐시 {len(cache)}건)…")
    rows = []
    for i, a in enumerate(apts, 1):
        # juso 조회: 시흥시 + 별표명 (실패 시 법정동 덧붙여 재시도)
        r = juso(f"시흥시 {a['apt']}", cache)
        if not r["juso"]:
            r = juso(f"시흥시 {a['dong']} {a['apt']}", cache)
        kb = known.get((a["code"], a["nn"]), set())
        cands = []                                   # juso 후보(파싱·정리)
        for j in r["juso"]:
            jn = nn(j["bdNm"])
            if not jn:
                continue
            m = JIBUN_IN_ADDR.search(j["jibun"])
            jji = m.group(2) if m else ""
            jibun_ok = bool(jji and jji in kb)
            if NONRES.search(j["bdNm"]) and not NONRES.search(a["apt"]) and not jibun_ok:
                continue                              # 비주거는 지번 확실할 때만
            cands.append({"bdNm": j["bdNm"], "jn": jn, "jji": jji,
                          "jibun_ok": jibun_ok})
        # 1순위: 지번이 별표/override와 일치하는 후보(이름 달라도 같은 단지) = 확정
        matched = _dedup([c for c in cands if c["jibun_ok"] and c["jn"] != a["nn"]])
        # 2순위: 지번 못 맞춤 → 이름 포함/확장 후보(4글자+, 변별력)
        sup = _dedup([c for c in cands
                      if len(a["nn"]) >= 4 and c["jn"] != a["nn"]
                      and (a["nn"] in c["jn"] or c["jn"] in a["nn"])])
        same = any(c["jn"] == a["nn"] for c in cands)
        if matched:
            aliases, basis, stat, why = matched, "지번일치", "O", ""
        elif len(sup) == 1:
            aliases, basis, stat, why = sup, "이름확장(단일후보)", "O", ""
        elif len(sup) >= 2:                           # 이름 공유 단지 여러 개 → 사람 확인
            aliases, basis, stat = sup, "이름확장(다수후보)", "X"
            why = "이름 공유 단지 여러 개 — 지번 확인 필요"
        elif same:
            aliases, basis, stat, why = [], "", "-", "별표명과 동일"
        elif not cands:
            aliases, basis, stat, why = [], "", "X", "juso 결과 없음"
        else:
            best = max(cands, key=lambda c: bigram(a["nn"], c["jn"]))
            aliases, basis, stat = [best], "유사후보", "X"
            why = f"이름·지번 불일치(유사도 {round(bigram(a['nn'], best['jn']),2)})"
        rows.append({**a, "aliases": aliases, "cands": cands,
                     "근거": basis, "자동확정": stat, "사유": why})
        if i % 40 == 0:
            json.dump(cache, open(CACHE, "w", encoding="utf-8"), ensure_ascii=False)
            print(f"  …{i}/{len(apts)}")
    json.dump(cache, open(CACHE, "w", encoding="utf-8"), ensure_ascii=False)

    # 자동확정 별칭 CSV — 확정(O)인 각 후보를 한 줄씩
    with open(ALIAS_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["code", "법정동(동단위)", "별표명", "별칭(juso공식명)", "juso지번", "근거"])
        n_alias = 0
        for r in rows:
            if r["자동확정"] != "O":
                continue
            for c in r["aliases"]:
                w.writerow([r["code"], r["dong"], r["apt"], c["bdNm"], c["jji"], r["근거"]])
                n_alias += 1
    n_auto = sum(1 for r in rows if r["자동확정"] == "O")
    print(f"\n자동확정 {n_auto}개 아파트 · 별칭 {n_alias}건 → {ALIAS_CSV}")

    write_xlsx(rows)
    n_same = sum(1 for r in rows if r["사유"] == "별표명과 동일")
    n_none = sum(1 for r in rows if r["사유"] == "juso 결과 없음")
    n_rev = sum(1 for r in rows if r["자동확정"] == "X")
    print(f"전체 {len(rows)} = 자동확정 {n_auto} / 동일 {n_same} / "
          f"juso없음 {n_none} / 검토필요 {n_rev}")
    print(f"검토 엑셀 → {XLSX}")


def _dedup(cands):
    """juso 정규화명(jn) 기준 중복 제거(첫 후보 유지)."""
    out, seen = [], set()
    for c in cands:
        if c["jn"] in seen:
            continue
        seen.add(c["jn"]); out.append(c)
    return out


def build_known():
    """(code, nn아파트) → set(지번): 별표 본문 + override 에서 검증된 지번 수집."""
    import gen_apartment_index as gi
    known = {}
    legmap = {}
    for cfgp in sorted(glob.glob("data/3*/config.json")):
        cfg = json.load(open(cfgp, encoding="utf-8"))
        legmap[cfg["admin_code"]] = (cfg["admin_dong"], list(cfg["legal_dongs"].values()))
    jibre = re.compile(r"([가-힣]{2,4}[동리])\s*(산?\d+(?:-\d+)?)")
    for cfgp in sorted(glob.glob("data/3*/config.json")):
        cfg = json.load(open(cfgp, encoding="utf-8"))
        code = cfg["admin_code"]
        csvp = f"data/{code}/관할구역.csv"
        if code not in legmap or not os.path.exists(csvp):
            continue
        beops = legmap[code][1]
        for r in csv.DictReader(open(csvp, encoding="utf-8-sig")):
            txt = str(r.get("관할구역", "")).strip()
            firsts = [mm.start() for rx in (gi.DONG_TOK, gi.DONG_KOR, gi.DONG_ALPHA)
                      for mm in [rx.search(txt)] if mm]
            if not firsts:
                continue
            head = txt[:min(firsts)]
            name = gi.apt_name(head, beops)
            if not name:
                continue
            k = (code, nn(name))
            for bdong, ji in jibre.findall(head):
                if bdong in beops:
                    known.setdefault(k, set()).add(ji)
    ovp = "data/apt_jibun_overrides.csv"
    if os.path.exists(ovp):
        for o in csv.DictReader(open(ovp, encoding="utf-8-sig")):
            k = (str(o["code"]), nn(o.get("아파트", "")))
            known.setdefault(k, set()).add(o["지번"].strip())
    return known


def bigram(a, b):
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    A = [a[i:i+2] for i in range(len(a)-1)]
    B = set(b[i:i+2] for i in range(len(b)-1))
    if not A or not B:
        return 0.0
    inter = sum(1 for x in A if x in B)
    return 2*inter/(len(A)+len(B))


def write_xlsx(rows):
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "별칭검토"
    head = ["동", "별표명", "juso 공식 건물명(별칭)", "juso지번",
            "근거", "자동확정", "채택(O/X)", "직접입력 별칭", "비고", "juso 후보 전체"]
    ws.append(head)
    bold = Font(bold=True)
    yel = PatternFill("solid", fgColor="FFF2CC")
    grn = PatternFill("solid", fgColor="E2EFDA")
    for c in ws[1]:
        c.font = bold
        c.alignment = Alignment(horizontal="center")
    # 검토 필요(불일치) 먼저, 그다음 자동확정, 동일/없음 마지막
    def order(r):
        if r["자동확정"] == "X" and r["사유"] not in ("별표명과 동일", "juso 결과 없음"):
            return 0
        if r["자동확정"] == "O":
            return 1
        return 2
    for r in sorted(rows, key=lambda r: (order(r), r["dong"], r["apt"])):
        adopt = "O" if r["자동확정"] == "O" else ""
        al = r["aliases"]
        names = " / ".join(c["bdNm"] for c in al)
        jibs = " / ".join(c["jji"] for c in al if c["jji"])
        allc = "  ·  ".join(f"{c['bdNm']}({c['jji']})" for c in r["cands"][:6])
        ws.append([r["dong"], r["apt"], names, jibs,
                   r["근거"], r["자동확정"], adopt, "", r["사유"], allc])
        last = ws[ws.max_row]
        if order(r) == 0:
            for c in last:
                c.fill = yel
        elif r["자동확정"] == "O":
            for c in last:
                c.fill = grn
    if ws.max_row >= 2:
        dv = DataValidation(type="list", formula1='"O,X"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"G2:G{ws.max_row}")
    widths = [10, 26, 32, 12, 16, 9, 10, 22, 24, 50]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = wdt
    ws.freeze_panes = "A2"
    # 안내 시트
    ws2 = wb.create_sheet("설명")
    for line in [
        ["별표↔실제명 별칭 검토표"],
        [""],
        ["문제", "별표는 축약명, 실제·공식명은 더 김 → 실제명으로 검색하면 안 나옴"],
        ["예시", "별표 '이편한세상 시흥장현' = 실제 'e편한세상 시흥장현 퍼스트베뉴'"],
        [""],
        ["노란칸", "이름·지번이 안 맞아 사람이 확인해야 함 → '채택(O/X)'에 직접 표시"],
        ["초록칸", "juso명이 별표명을 포함/확장하거나 지번 일치 → 자동확정(O)"],
        ["채택 O", "이 juso명을 검색 별칭으로 등록(실제명으로도 검색됨)"],
        ["채택 X", "juso명이 틀림 → 등록 안 함. 맞는 실제명을 알면 '직접입력 별칭'에 기입"],
        [""],
        ["반영", "검토 후 이 파일을 알려주시면 apt_aliases.csv에 합쳐 재빌드합니다"],
    ]:
        ws2.append(line)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 70
    wb.save(XLSX)


if __name__ == "__main__":
    main()
