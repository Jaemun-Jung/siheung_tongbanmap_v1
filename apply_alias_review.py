#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
별표_별칭_검토.xlsx 의 담당자 확정분을 data/apt_aliases.csv 에 추가.

규칙: 자동확정이 아니었던(노란) 행 중 담당자가
  - 채택(O/X)='O' 로 표시했거나
  - '직접입력 별칭' 을 적은 행
만 추가한다(직접입력이 있으면 그 이름을 우선). juso 무응답·별표명 동일·미표시(X)는 건너뜀.
중복(같은 code+별표명+별칭)은 추가하지 않음.
사용: PYTHONUTF8=1 python apply_alias_review.py
"""
import csv, glob, json, os
from openpyxl import load_workbook

XLSX = "out/_audit/별표_별칭_검토.xlsx"
CSV = "data/apt_aliases.csv"


def main():
    # 동 이름 → admin_code
    name2code = {}
    for cfgp in sorted(glob.glob("data/3*/config.json")):
        cfg = json.load(open(cfgp, encoding="utf-8"))
        name2code[cfg["admin_dong"]] = cfg["admin_code"]

    wb = load_workbook(XLSX)
    ws = wb["별칭검토"]
    H = {c.value: i + 1 for i, c in enumerate(ws[1])}

    def g(r, name):
        v = ws.cell(r, H[name]).value
        return "" if v is None else str(v).strip()

    add = []
    for r in range(2, ws.max_row + 1):
        if g(r, "자동확정") == "O":            # 이미 반영된 초록 행은 건너뜀
            continue
        take = g(r, "채택(O/X)")
        manual = g(r, "직접입력 별칭")
        if take != "O" and manual == "":        # 담당자가 채택 안 함
            continue
        dong = g(r, "동")
        code = name2code.get(dong)
        if not code:
            print(f"  ⚠ {r}행 동코드 없음: {dong}"); continue
        pyo = g(r, "별표명")
        alias = manual or g(r, "juso 공식 건물명(별칭)")
        jibun = g(r, "juso지번") if not manual else ""   # 직접입력은 juso지번이 무의미
        basis = "담당자 직접입력" if manual else "담당자 확인(juso)"
        add.append([code, dong, pyo, alias, jibun, basis])

    # 기존 CSV 로드 + 중복 제거
    rows = list(csv.reader(open(CSV, encoding="utf-8-sig")))
    header, body = rows[0], rows[1:]
    seen = {(r[0], r[2], r[3]) for r in body}
    new = 0
    for a in add:
        if (a[0], a[2], a[3]) in seen:
            print(f"  · 이미 있음: {a[1]} {a[2]} → {a[3]}"); continue
        body.append(a); seen.add((a[0], a[2], a[3])); new += 1
        print(f"  + {a[1]} {a[2]} → {a[3]}  [{a[5]}]")

    with open(CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(body)
    print(f"\n→ 신규 별칭 {new}건 추가, 총 {len(body)}건 → {CSV}")


if __name__ == "__main__":
    main()
