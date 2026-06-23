#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
정합성 검사 결과(out/_audit/정합성_불일치.csv) → 담당자 검토 엑셀.
각 누락 지번에 '지목'(대지/도로/전답 등)과 '추정유형'을 붙여 검토 우선순위를 돕는다.
대지(집 짓는 땅) 누락이 가장 중요(별표 누락 의심), 도로·전답은 통상 정상.
사용: PYTHONUTF8=1 python gen_verify_excel.py   (verify_byeolpyo.py 먼저 실행)
"""
import csv, glob, json, os, re
import geopandas as gpd

CSV = "out/_audit/정합성_불일치.csv"
XLSX = "out/_audit/정합성_검토.xlsx"
SHP = "LSMD_CONT_LDREG_41390_202606.shp"


def canon(pnu):
    pnu = str(pnu)
    if len(pnu) < 19 or not pnu[:19].isdigit():
        return None, None
    return pnu[:10], ('산' if pnu[10] == '2' else '') + str(int(pnu[11:15])) + \
        (f'-{int(pnu[15:19])}' if int(pnu[15:19]) else '')


def jimok(j):
    m = re.match(r'^산?\d[\d-]*(\D*)$', str(j))
    return m.group(1).strip() if m else ''


def is_range_artifact(j):
    # '1773-1774', '400-499' 처럼 '본번-본번'(부번이 본번보다 큰) = 파서 범위 추정
    m = re.fullmatch(r'(\d+)-(\d+)', str(j))
    return bool(m) and len(m.group(2)) >= 3 and int(m.group(2)) > int(m.group(1))


def main():
    if not os.path.exists(CSV):
        raise SystemExit("먼저 verify_byeolpyo.py 를 실행하세요.")
    rows = list(csv.DictReader(open(CSV, encoding="utf-8-sig")))
    # 지목 사전 (법정동명, 지번) -> 지목
    legname = {}
    for cfgp in glob.glob("data/3*/config.json"):
        legname.update(json.load(open(cfgp, encoding="utf-8"))["legal_dongs"])
    shp = gpd.read_file(SHP)
    shp["_bjd"], shp["_jib"] = zip(*shp["PNU"].map(canon))
    jm = {}
    for _, r in shp[shp["_bjd"].isin(legname.keys())].iterrows():
        jm[(legname[r["_bjd"]], r["_jib"])] = jimok(r["JIBUN"])

    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "누락검토"
    head = ["행정동", "법정동", "지번", "별표통", "지목", "추정유형", "검토(O/X)", "비고"]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
    yel = PatternFill("solid", fgColor="FFF2CC")    # 대지 = 중요
    gry = PatternFill("solid", fgColor="EFEFEF")    # 파서artifact/비대지

    drops = [r for r in rows if r["유형"] == "누락"]

    def sortkey(r):
        z = jm.get((r["법정동"], r["지번"]), "")
        art = is_range_artifact(r["지번"])
        # 대지(0) > 기타지목(1) > 파서artifact(2)
        return (2 if art else (0 if z.startswith("대") else 1), r["행정동"], r["지번"])

    for r in sorted(drops, key=sortkey):
        z = jm.get((r["법정동"], r["지번"]), "?")
        if is_range_artifact(r["지번"]):
            typ = "파서 범위 추정(무시 가능)"
        elif z.startswith("대"):
            typ = "★ 대지 — 별표 누락 의심(중요)"
        else:
            typ = "비대지(도로·전답 등 — 대개 정상)"
        m = re.search(r"별표 \[?(\d[\d, ]*)", r["상세"])
        byp = m.group(1) if m else ""
        ws.append([r["행정동"], r["법정동"], r["지번"], byp, z, typ, "", ""])
        last = ws[ws.max_row]
        if z.startswith("대") and not is_range_artifact(r["지번"]):
            for c in last:
                c.fill = yel
        elif is_range_artifact(r["지번"]):
            for c in last:
                c.fill = gry

    if ws.max_row >= 2:
        dv = DataValidation(type="list", formula1='"O,X"', allow_blank=True)
        ws.add_data_validation(dv); dv.add(f"G2:G{ws.max_row}")
    for i, w in enumerate([9, 9, 12, 8, 7, 30, 10, 24], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    # 참고 시트: 오배정(검증된 override) + 설명
    ws2 = wb.create_sheet("참고_오배정")
    ws2.append(["행정동", "법정동", "지번", "내용"])
    for r in [r for r in rows if r["유형"] == "오배정"]:
        ws2.append([r["행정동"], r["법정동"], r["지번"], r["상세"] + " (담당자 검증 override — 정상)"])
    ws3 = wb.create_sheet("설명")
    for line in [
        ["정합성 검토표 — 별표엔 있는데 지도에 안 그려진 지번"],
        [""],
        ["노란칸 ★", "지목이 '대지'인데 통에 안 들어감 → 별표 누락 의심. 진짜 빠진 거면 검토(O)에 O"],
        ["회색칸", "지번이 범위처럼 보임(예 1773-1774) = 파서 추정, 대개 무시 가능"],
        ["흰칸", "도로·전답·임야 등 비대지 → 통에 안 들어가도 통상 정상(참고용)"],
        [""],
        ["검토 O", "진짜로 빠졌고 통에 들어가야 함 → 수정 대상(지번·통 알려주시면 반영)"],
        ["검토 X", "정상(폐번·비대지·범위표기 등) → 둘 것"],
        [""],
        ["참고_오배정 시트", "지도가 별표와 다른 2곳 = 담당자가 juso로 검증해 일부러 고친 곳(정상)"],
    ]:
        ws3.append(line)
    ws3.column_dimensions["A"].width = 16; ws3.column_dimensions["B"].width = 70

    wb.save(XLSX)
    n_dae = sum(1 for r in drops if jm.get((r["법정동"], r["지번"]), "").startswith("대")
                and not is_range_artifact(r["지번"]))
    print(f"누락 {len(drops)}건 → {XLSX}")
    print(f"  ★ 대지(중요) {n_dae} · 파서artifact {sum(1 for r in drops if is_range_artifact(r['지번']))} · 나머지 비대지")


if __name__ == "__main__":
    main()
